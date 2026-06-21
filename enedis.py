"""Lecteur d'export Enedis (conso quotidienne kWh).

Parse les .xlsx "Export_energie_Consommation_*.xlsx" et renvoie une série
[(date, kWh)] + des stats. Réutilisable d'un mois à l'autre, et destiné à
être croisé avec climate_log.csv (météo/clim/gaming).

Usage:  uv run python enedis.py <fichier.xlsx>
"""
from __future__ import annotations

import datetime as dt
import statistics as stats
import sys
from pathlib import Path

import openpyxl

PRICE_EUR_PER_KWH = 0.25  # tarif indicatif; à ajuster (HP/HC, Tempo...)


def load(path: str) -> list[tuple[dt.date, float]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = next((w for w in wb.worksheets if "Consommation" in w.title), wb.worksheets[-1])
    out = []
    for row in ws.iter_rows(values_only=True):
        d, v = row[1], row[2]
        if isinstance(v, (int, float)) and isinstance(d, str) and "/" in d:
            out.append((dt.datetime.strptime(d, "%d/%m/%Y").date(), float(v)))
    out.sort()
    return out


def summary(data: list[tuple[dt.date, float]]) -> dict:
    vals = [v for _, v in data]
    base = min(vals)
    wk = [v for d, v in data if d.weekday() < 5]
    we = [v for d, v in data if d.weekday() >= 5]
    return {
        "days": len(data), "start": data[0][0], "end": data[-1][0],
        "total_kwh": sum(vals), "mean_kwh": stats.mean(vals),
        "baseline_kwh": base, "baseline_w": base / 24 * 1000,
        "above_baseline_kwh": sum(v - base for v in vals),
        "weekday_mean": stats.mean(wk) if wk else None,
        "weekend_mean": stats.mean(we) if we else None,
        "cost_eur": sum(vals) * PRICE_EUR_PER_KWH,
    }


def load_hourly(path: str) -> list[tuple[dt.datetime, float, str]]:
    """Courbe de charge / conso horaire : lignes horodatées (date+heure).
    Renvoie [(datetime, valeur, unite)]. Auto-détecte W vs kWh via l'entête.
    NB: format à confirmer sur le vrai fichier (colonnes variables selon export).
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = next((w for w in wb.worksheets if "Consommation" in w.title or "charge" in w.title.lower()),
              wb.worksheets[-1])
    unit = "kWh"
    out = []
    for row in ws.iter_rows(values_only=True):
        cells = [c for c in row if c is not None]
        # repère l'unité dans l'entête
        for c in cells:
            if isinstance(c, str) and "(en W" in c:
                unit = "W"
            elif isinstance(c, str) and "(en kWh" in c:
                unit = "kWh"
        ts = val = None
        for c in cells:
            if isinstance(c, dt.datetime):
                ts = c
            elif isinstance(c, str) and "/" in c and ":" in c:
                for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M"):
                    try:
                        ts = dt.datetime.strptime(c, fmt); break
                    except ValueError:
                        pass
            elif isinstance(c, (int, float)):
                val = float(c)
        if ts is not None and val is not None:
            out.append((ts, val, unit))
    out.sort(key=lambda x: x[0])
    return out


def night_floor(hourly: list[tuple[dt.datetime, float, str]],
                start_h: int = 2, end_h: int = 5) -> dict:
    """Plancher nocturne = vrai talon. Conso 2h-5h, convertie en W moyens."""
    night = [(t, v, u) for t, v, u in hourly if start_h <= t.hour < end_h]
    if not night:
        return {}
    unit = night[0][2]
    # en W si unit==W; si kWh par pas de 30min -> W = kWh*2*1000 ; par heure -> *1000
    def to_w(v, u):
        return v if u == "W" else v * 1000  # affine selon le pas réel du fichier
    watts = [to_w(v, u) for _, v, u in night]
    watts.sort()
    return {"samples": len(watts), "floor_w_min": watts[0],
            "floor_w_median": watts[len(watts) // 2],
            "note": "si kWh, ajuster x2 selon pas 30min vs 60min"}


def _parse_dt(v):
    if isinstance(v, dt.datetime):
        return v
    if isinstance(v, str):
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M"):
            try:
                return dt.datetime.strptime(v, fmt)
            except ValueError:
                pass
    return None


def load_curve_from_export(path: str) -> list[tuple[dt.datetime, float]]:
    """Lit la feuille 'Courbe de charge 30min' produite par enedis_export.py
    (colonnes: Date début | Date fin | Puissance kW). Renvoie [(datetime, W)]."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = next((s for s in wb.sheetnames if "ourbe" in s), None)
    if not sheet:
        return []
    out = []
    for row in wb[sheet].iter_rows(min_row=2, values_only=True):
        ts, kw = _parse_dt(row[0]), row[2]
        if ts is not None and isinstance(kw, (int, float)):
            out.append((ts, float(kw) * 1000))  # kW -> W
    out.sort()
    return out


def floor_w(curve: list[tuple[dt.datetime, float]], start_h=2, end_h=5) -> dict:
    """Plancher nocturne (W) = vrai talon de la maison."""
    night = sorted(v for t, v in curve if start_h <= t.hour < end_h)
    if not night:
        return {"note": "aucun point nuit (données pas encore dispo ?)"}
    return {"samples": len(night), "floor_w_min": round(night[0]),
            "floor_w_median": round(night[len(night) // 2]),
            "floor_kwh_per_day_equiv": round(night[len(night) // 2] * 24 / 1000, 1)}


if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else None
    if not path or not Path(path).exists():
        sys.exit("Usage: uv run python enedis.py <export.xlsx>  [--hourly|--export-curve]")
    if "--export-curve" in sys.argv:
        c = load_curve_from_export(path)
        if not c:
            sys.exit("Pas de feuille 'Courbe de charge 30min' (lancer enedis_export.py --courbe).")
        print(f"{len(c)} points 30min ({c[0][0]} -> {c[-1][0]})")
        print("Plancher nuit 2h-5h:", floor_w(c))
        sys.exit(0)
    if "--hourly" in sys.argv:
        h = load_hourly(path)
        print(f"{len(h)} points horodatés ({h[0][0]} -> {h[-1][0]}), unité={h[0][2]}")
        print("Plancher nuit (2h-5h):", night_floor(h))
    else:
        data = load(path)
        s = summary(data)
        print(f"{s['days']} jours ({s['start']} -> {s['end']})")
        print(f"Total {s['total_kwh']:.1f} kWh | moy {s['mean_kwh']:.1f} kWh/j | ~{s['cost_eur']:.0f} €")
        print(f"Baseline {s['baseline_kwh']:.1f} kWh/j (~{s['baseline_w']:.0f} W permanents)")
        print(f"Semaine {s['weekday_mean']:.1f} | Week-end {s['weekend_mean']:.1f} kWh/j")
