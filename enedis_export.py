#!/usr/bin/env python3
"""
Export de la consommation Enedis en .xlsx — sans API tierce.

Deux types de données, même endpoint interne du portail (pas une API tierce) :

    https://alex.microapplications.enedis.fr/mes-mesures-prm/api/private/v2/
        personnes/{personneId}/prms/{prm}/donnees-energetiques

  • ENERGIE  (défaut) : agrégats jour / semaine / mois / année, kWh, ~3 ans.
  • PUISSANCE (--courbe) : courbe de charge au pas demi-horaire, kW.
        ⚠ Nécessite d'avoir ACTIVÉ "l'enregistrement de la consommation au
        pas demi-horaire" sur le portail Enedis. Sinon l'API renvoie 401
        ("Veuillez activer les données depuis Visualiser vos mesures").
        La courbe se récupère par fenêtres de dates (le script boucle par 7 j).

Authentification : cookie de session, via un Chromium piloté par Playwright
avec PROFIL PERSISTANT (connexion manuelle une seule fois).

Installation (déjà fait dans ce projet uv)
------------
    uv add playwright openpyxl
    uv run playwright install chromium

Utilisation
-----------
    uv run python enedis_export.py                           # agrégats -> conso_enedis.xlsx
    uv run python enedis_export.py --courbe                  # + courbe de charge (7 derniers jours)
    uv run python enedis_export.py --courbe --debut 2026-05-01 --fin 2026-06-19
    uv run python enedis_export.py -o sortie.xlsx --headless # après la 1re connexion

Puis analyse du plancher nocturne :
    uv run python enedis.py conso_enedis.xlsx --export-curve
"""

import argparse
import sys
from datetime import date, datetime, timedelta
from pathlib import Path
from urllib.parse import urlparse, parse_qs

PAGE_URL = "https://mon-compte-particulier.enedis.fr/visualiser-vos-mesures-consommation"
API_FRAGMENT = "/donnees-energetiques"
PROFILE_DIR = Path.home() / ".enedis_export_profile"
WINDOW_DAYS = 7  # la courbe de charge se récupère par fenêtres courtes

AGG_LABELS = {
    "jour": "Journalier",
    "semaine": "Hebdomadaire",
    "mois": "Mensuel",
    "annee": "Annuel",
}


def to_float(v):
    if v is None:
        return None
    s = str(v).strip().replace(" ", "").replace(" ", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def daterange_windows(d_start, d_end, step):
    cur = d_start
    while cur <= d_end:
        end = min(cur + timedelta(days=step - 1), d_end)
        yield cur.isoformat(), (end + timedelta(days=1)).isoformat()  # fin exclusive
        cur = end + timedelta(days=1)


def run(args):
    from playwright.sync_api import sync_playwright

    with sync_playwright() as pw:
        ctx = pw.chromium.launch_persistent_context(
            user_data_dir=str(PROFILE_DIR),
            headless=args.headless,
            args=["--disable-blink-features=AutomationControlled"],
        )
        page = ctx.pages[0] if ctx.pages else ctx.new_page()

        captured = {}
        page.on("request", lambda r: captured.__setitem__("url", r.url)
                if (API_FRAGMENT in r.url and "personnes" in r.url) else None)

        print("→ Ouverture du portail Enedis…")
        page.goto(PAGE_URL, wait_until="domcontentloaded")
        print("  Connecte-toi dans la fenêtre si besoin. J'attends les données (max 5 min)…")
        try:
            page.wait_for_request(
                lambda r: API_FRAGMENT in r.url and "personnes" in r.url, timeout=300_000)
        except Exception:
            ctx.close()
            sys.exit("✗ Aucune requête de données détectée (connexion incomplète ?).")

        api_url = captured["url"]
        base = api_url.split(API_FRAGMENT)[0]            # .../prms/{prm}
        q = parse_qs(urlparse(api_url).query)
        segments = (q.get("segments") or ["C5"])[0]
        typed = (q.get("typeDonnees") or ["CONS"])[0]
        print("→ Endpoint détecté (personne + PRM identifiés automatiquement).")

        # ---- 1) Agrégats ENERGIE -------------------------------------------
        energie = page.evaluate(
            """async ({base, seg, td}) => {
                const url = base + "/donnees-energetiques?mesuresTypeCode=ENERGIE"
                  + "&mesuresCorrigees=false&typeDonnees=" + td + "&segments=" + seg;
                const r = await fetch(url, {credentials:'include'});
                if (!r.ok) return {error: r.status};
                return await r.json();
            }""",
            {"base": base, "seg": segments, "td": typed},
        )

        # ---- 2) Courbe de charge PUISSANCE (optionnel) ---------------------
        courbe = None
        if args.courbe:
            d_end = datetime.strptime(args.fin, "%Y-%m-%d").date() if args.fin else date.today() - timedelta(days=1)
            d_start = datetime.strptime(args.debut, "%Y-%m-%d").date() if args.debut else d_end - timedelta(days=6)
            windows = list(daterange_windows(d_start, d_end, WINDOW_DAYS))
            urls = [
                base + "/donnees-energetiques?mesuresTypeCode=PUISSANCE"
                + "&mesuresCorrigees=false&typeDonnees=" + typed
                + "&segments=" + segments + "&dateDebut=" + a + "&dateFin=" + b
                for (a, b) in windows
            ]
            print(f"→ Courbe de charge : {len(urls)} fenêtre(s) de {WINDOW_DAYS} j de {d_start} à {d_end}…")
            courbe = page.evaluate(
                """async (urls) => {
                    const pts = [];
                    for (const u of urls) {
                        const r = await fetch(u, {credentials:'include'});
                        if (r.status === 401) return {error: 401};
                        if (!r.ok) continue;
                        const j = await r.json();
                        const root = j.cons || j.prod || j;
                        // cible précise : tableau dont les éléments sont des POINTS
                        // (ont 'valeur' + un champ de date). Évite d'attraper un
                        // tableau quelconque.
                        const isPointArr = (v) => Array.isArray(v) && v.length
                            && v[0] && typeof v[0] === 'object' && 'valeur' in v[0]
                            && ('dateDebut' in v[0] || 'dateHeure' in v[0] || 'date' in v[0]);
                        const find = (o) => {
                            if (isPointArr(o)) return o;
                            if (o && typeof o === 'object') {
                                for (const k in o) { const r = find(o[k]); if (r) return r; }
                            }
                            return null;
                        };
                        const arr = find(root);
                        if (arr) pts.push(...arr);
                    }
                    return {points: pts};
                }""",
                urls,
            )

        ctx.close()
        return energie, courbe


def write_xlsx(energie, courbe, out_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)
    hf = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="1F4E78")

    def style_header(ws, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=1, column=c)
            cell.font, cell.fill = hf, fill
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"

    total = 0
    aggregats = ((energie or {}).get("cons") or {}).get("aggregats") or {}
    for key, label in AGG_LABELS.items():
        donnees = (aggregats.get(key) or {}).get("donnees") or []
        ws = wb.create_sheet(title=label)
        ws.append(["Date début", "Date fin", "Consommation (kWh)"])
        for d in donnees:
            ws.append([(d.get("dateDebut") or "").replace("T", " "),
                       (d.get("dateFin") or "").replace("T", " "),
                       to_float(d.get("valeur"))])
        for i, w in enumerate([22, 26, 20], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
            row[0].number_format = "0.00"
        style_header(ws, 3)
        total += len(donnees)

    if courbe and not courbe.get("error"):
        pts = courbe.get("points") or []
        ws = wb.create_sheet(title="Courbe de charge 30min")
        ws.append(["Date début", "Date fin", "Puissance (kW)"])
        for d in pts:
            ws.append([(d.get("dateDebut") or "").replace("T", " "),
                       (d.get("dateFin") or "").replace("T", " "),
                       to_float(d.get("valeur"))])
        for i, w in enumerate([22, 26, 18], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
            row[0].number_format = "0.000"
        style_header(ws, 3)
        total += len(pts)

    wb.save(out_path)
    return total


def main():
    ap = argparse.ArgumentParser(description="Export conso Enedis en xlsx (sans API tierce).")
    ap.add_argument("-o", "--output", default="conso_enedis.xlsx")
    ap.add_argument("--courbe", action="store_true", help="Ajouter la courbe de charge au pas 1/2 h.")
    ap.add_argument("--debut", help="Début courbe de charge (YYYY-MM-DD).")
    ap.add_argument("--fin", help="Fin courbe de charge (YYYY-MM-DD).")
    ap.add_argument("--headless", action="store_true", help="Sans fenêtre (après la 1re connexion).")
    args = ap.parse_args()

    energie, courbe = run(args)

    if energie and energie.get("error"):
        sys.exit(f"✗ Données agrégées indisponibles (HTTP {energie['error']}).")
    if args.courbe and courbe and courbe.get("error") == 401:
        print("\n⚠ Courbe de charge indisponible (HTTP 401).")
        print("  La collecte au pas demi-horaire n'est pas activée sur ce compte.")
        print("  Active-la sur le portail : « Visualiser vos mesures » → activer")
        print("  l'enregistrement de la consommation au pas demi-horaire, puis réessaie")
        print("  (les données ne sont disponibles qu'à partir de l'activation).")

    out_path = Path(args.output).expanduser().resolve()
    n = write_xlsx(energie, courbe, out_path)
    print(f"\n✓ {n} lignes écrites dans {out_path}")


if __name__ == "__main__":
    main()
